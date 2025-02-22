print("===)]==============--->"),
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.preprocessing import LabelEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import GaussianNB
from sklearn.model_selection import train_test_split
import tkinter as tk
from matplotlib.figure import Figure 
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg,  
NavigationToolbar2Tk) 
from openpyxl import load_workbook

veriler = pd.read_excel("Book.xlsx")
print(veriler)
# erkek_veriler = veriler[veriler["cinsiyet"].isin(["e"])]
# erkek_yaslar = erkek_veriler.iloc[:,0:1].values
# np_erkek_yaslar = np.array(erkek_yaslar)

# erkek_kilo = erkek_veriler.iloc[:,1:2]
# np_erkek_kilo = np.array(erkek_kilo)

# plt.scatter(np_erkek_yaslar,np_erkek_kilo)
# plt.show()

x_ = veriler.iloc[:,0:3].values
y = veriler.iloc[:,3:].values


# Kategorik değişkenleri sayısala çevir
label_encoder = LabelEncoder()
y_ = label_encoder.fit_transform(y)

x_train, x_test, y_train, y_test = train_test_split(x_, y_, test_size=0.2, random_state=42)

# random forest Modeli oluştur ve eğit
rf = RandomForestClassifier(n_estimators=100, random_state=42)
rf.fit(x_train, y_train)
 
# LogisticRegression modeli oluştur ve eğit
logr = LogisticRegression(random_state=0)
logr.fit(x_train,y_train) 
 
# Gaussian modeli oluştur ve eğit
gnb = GaussianNB()
gnb.fit(x_train, y_train)
 
root = tk.Tk()
root.title("Cinsiyet Tahmini")
root.geometry("500x600")
var  = tk.StringVar()
var1 = tk.StringVar()
var2 = tk.StringVar()
pre_list=[]

tk.Label(root, text="Yaşınızı girin:").pack()
yas_entry = tk.Entry(root)
yas_entry.pack()

tk.Label(root, text="Boyunuzu (cm) girin:").pack()
boy_entry = tk.Entry(root)
boy_entry.pack()

tk.Label(root, text="Kilonuzu (kg) girin:").pack()
kilo_entry = tk.Entry(root)
kilo_entry.pack()

cinsiyet_var = tk.IntVar()
frame = tk.Frame(root)
frame.pack()

tk.Radiobutton(frame, text="e", variable=cinsiyet_var, value=1).pack(side=tk.LEFT)
tk.Radiobutton(frame, text="k", variable=cinsiyet_var, value=0).pack(side=tk.LEFT)

def tahmin_et():
   
    yas = int(yas_entry.get())
    boy = int(boy_entry.get())
    kilo = int(kilo_entry.get())
    gercek_cinsiyet = "e" if cinsiyet_var.get() == 1 else "k"
    
    pre_rf=rf.predict([[yas,kilo,boy]])
    pre_list.append(pre_rf)
    pre_lr=logr.predict([[yas,kilo,boy]])
    pre_list.append(pre_lr)
    pre_gaus=gnb.predict([[yas,kilo,boy]])
    pre_list.append(pre_gaus)
    
    yeni_veri = pd.DataFrame([[yas, kilo, boy, gercek_cinsiyet]], columns=['yas', 'Kilo', 'boy', 'cinsiyet'])
    excel_path = "Book.xlsx"
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            yeni_veri.to_excel(writer, index=False, header=False, startrow=sheet.max_row)
    except FileNotFoundError:    
            yeni_veri.to_excel(excel_path, index=False, engine='openpyxl')
    
    for i in pre_list :
        if(i):
            print("cinsiyetiniz kız")
            var.set("Random Forest modeli cinsiyetinizi kız tahmin etti")
            var1.set("Lineer Regression modeli cinsiyetinizi kız tahmin etti")
            var2.set("Gaussian modeli cinsiyetinizi kız tahmin etti")
        else:
           print("cinsiyetiniz erkek")
           var.set("Random Forest modeli cinsiyetiniz erkek tahmin etti")
           var1.set("Lineer Regression modeli cinsiyetiniz erkek tahmin etti")
           var2.set("Gaussian modeli cinsiyetiniz erkek tahmin etti")

def plot(): 
  
    # the figure that will contain the plot 
    fig = Figure(figsize = (4, 4), dpi = 100) 
    plot1 = fig.add_subplot(111)
    
    global erkek_veriler
    erkek_veriler = veriler[veriler["cinsiyet"].isin(["e"])]
    erkek_yaslar = erkek_veriler.iloc[:,0:1].values
    np_erkek_yaslar = np.array(erkek_yaslar)
    erkek_kilo = erkek_veriler.iloc[:,1:2]
    np_erkek_kilo = np.array(erkek_kilo)

    plot1.scatter(np_erkek_yaslar,np_erkek_kilo)
    # plot1.title("erkeklerin yaşlarına göre kilo dağılımları")
    # plot1.xlabel("erkek yas")
    # plot1.ylabel("erkek kilo")
    plt.show()
    # creating the Tkinter canvas 
    
    global kadin_veriler 
    kadin_veriler = veriler[veriler["cinsiyet"].isin(["k"])]
    kadin_yaslar = kadin_veriler.iloc[:,0:1].values
    np_kadin_yaslar = np.array(kadin_yaslar)
    kadin_kilo = kadin_veriler.iloc[:,1:2]
    np_kadin_kilo = np.array(kadin_kilo)
    plot1.scatter(np_kadin_yaslar,np_kadin_kilo,color='red')
    
    plt.show()
    
    # containing the Matplotlib figure 
    canvas = FigureCanvasTkAgg(fig, master = root)   
    canvas.draw() 
  
    # placing the canvas on the Tkinter window 
    canvas.get_tk_widget().pack() 
  
    # creating the Matplotlib toolbar 
    toolbar = NavigationToolbar2Tk(canvas, root) 
    toolbar.update() 
  
    # placing the toolbar on the Tkinter window 
    canvas.get_tk_widget().pack() 
tk.Button(root, text="Predict", command=tahmin_et).pack()
tk.Button(root, text="Show Plot", command=plot).pack()

tk.Label(textvariable=var).pack()
tk.Label(textvariable=var1).pack()
tk.Label(textvariable=var2).pack()
root.mainloop()
